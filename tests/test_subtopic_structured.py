"""Bounded agency-declared P9 child routes and their fail-closed canaries."""

from io import BytesIO
import sys
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook  # noqa: E402

from scripts import subtopic_records as records  # noqa: E402
from scripts import subtopic_structured as structured  # noqa: E402


def extractor(content, content_type, name, url):
    del content_type, name, url
    return ([{"page": 1, "text": content.decode("utf-8")}], {})


def attachment_boundary(name, content, *, url="https://files.example/source"):
    document = {
        "url": url,
        "name": name,
        "content_type": "application/octet-stream",
        "sha256": "fixture-hash",
        "source_kind": "authoritative_notice",
    }

    def detail(_opportunity_id):
        return {"data": {"attachments": True}}

    def collect(_data):
        return [{"download_url": url, "file_name": name, "id": "attachment-1"}]

    def download(_url):
        raise AssertionError("the supplied current attachment must not be refetched")

    return document, detail, collect, download


class HgeoTests(unittest.TestCase):
    def text(self):
        return "\n".join(
            [
                "===== 4 =====",
                "Subtopic 1A: Coal",
                "Applicants select this subtopic.",
                "Subtopic 1B: Oil & Gas",
                "Applicants select this subtopic.",
                "Subtopic 1C: Geothermal",
                "Applicants select this subtopic.",
            ]
        )

    def test_the_adjudicated_hgeo_parent_emits_exactly_three_children(self):
        content = self.text().encode()
        name = "FundOpp_DE-FOA-0003215.pdf"
        document, detail, collect, download = attachment_boundary(name, content)
        outcome = structured.first_refusal(
            {
                "opportunity_id": "363594",
                "opportunity_number": "DE-FOA-0003215",
                "status": "posted",
            },
            content,
            document,
            detail_fetcher=detail,
            collector=collect,
            download=download,
            extract_containers=extractor,
            as_of="2026-08-20",
        )
        self.assertTrue(outcome.claimed)
        self.assertEqual(len(outcome.records), 3)
        self.assertEqual([item["subtopic_code"] for item in outcome.records],
                         ["1A", "1B", "1C"])
        for item in outcome.records:
            self.assertEqual(item["subtopic_source"], records.INLINE)
            self.assertEqual(item["child_type"], "subject")
            self.assertEqual(item["source_role"], "authoritative_announcement")

    def test_a_recognized_parent_with_collapsed_structure_fails_closed(self):
        content = b"===== 4 =====\nSubtopic 1A: Coal\n"
        name = "FundOpp_DE-FOA-0003215.pdf"
        document, detail, collect, download = attachment_boundary(name, content)
        outcome = structured.first_refusal(
            {
                "opportunity_id": "363594",
                "opportunity_number": "DE-FOA-0003215",
                "status": "posted",
            },
            content,
            document,
            detail_fetcher=detail,
            collector=collect,
            download=download,
            extract_containers=extractor,
            as_of="2026-08-20",
        )
        self.assertTrue(outcome.claimed)
        self.assertEqual(outcome.records, ())
        self.assertEqual(outcome.reason, "structured_source_failed")


class ArlTests(unittest.TestCase):
    def test_ids_titles_pages_and_summary_boundaries_are_preserved(self):
        text = (
            "===== 1 =====\nTitle: Quantum Sensing ARL-BAA-0001\n"
            "Description: Sensors and systems.\n"
            "===== 2 =====\nTitle: Energy Storage ARL-BAA-0002\n"
            "Description: Batteries and materials."
        )
        children = structured.parse_arl_topics(text)
        self.assertEqual([item["code"] for item in children],
                         ["ARL-BAA-0001", "ARL-BAA-0002"])
        self.assertEqual([item["title"] for item in children],
                         ["Quantum Sensing", "Energy Storage"])
        self.assertEqual([item["page_start"] for item in children], [1, 2])
        self.assertNotIn("Energy Storage", children[0]["summary"])

    def test_duplicate_topic_ids_collapse_the_route(self):
        text = (
            "Title: First ARL-BAA-0001\nDescription: A.\n"
            "Title: Second ARL-BAA-0001\nDescription: B."
        )
        self.assertEqual(structured.parse_arl_topics(text), [])


def genesis_workbook():
    book = Workbook()
    summary = book.active
    summary.title = "Phase I Summary"
    summary["A1"] = "Focus Area Select from dropdown menu"
    focus = book.create_sheet("Focus Areas")
    row = 1
    # 21 source groups and exactly 98 selectable focus areas: five in the
    # first 14 groups and four in the remaining seven.
    for group in range(1, 22):
        count = 5 if group <= 14 else 4
        for offset in range(count):
            letter = chr(ord("A") + offset)
            focus.cell(row=row, column=1).value = (
                f"{group}-{letter} Challenge {group} | Focus {group}-{letter}"
            )
            row += 1
    output = BytesIO()
    book.save(output)
    return output.getvalue()


class GenesisTests(unittest.TestCase):
    def test_the_workbook_has_21_groups_and_98_selectable_focus_areas(self):
        children, diagnostics = structured.parse_genesis_workbook(genesis_workbook())
        self.assertEqual(diagnostics, {"challenge_groups": 21, "focus_areas": 98})
        self.assertEqual(len(children), 119)
        self.assertEqual(
            sum(item["code"].startswith("challenge-") for item in children), 21
        )

    def test_focus_children_point_to_their_stable_group_child(self):
        content = genesis_workbook()
        name = "Genesis Mission Phase I Application Template v2.xlsx"
        document, detail, collect, download = attachment_boundary(name, content)
        outcome = structured.first_refusal(
            {
                "opportunity_id": "361526",
                "opportunity_number": "DE-FOA-0003612",
                "status": "posted",
            },
            content,
            document,
            detail_fetcher=detail,
            collector=collect,
            download=download,
            extract_containers=extractor,
            as_of="2026-08-20",
        )
        self.assertTrue(outcome.claimed)
        self.assertEqual(len(outcome.records), 119)
        first_focus = outcome.records[21]
        self.assertEqual(first_focus["group_id"], "challenge-1")
        self.assertEqual(first_focus["parent_subtopic_id"],
                         "361526:challenge-1")
        self.assertEqual(first_focus["child_type"], "subject")
        self.assertEqual(first_focus["subtopic_source"], records.NATIVE)


class BindingTests(unittest.TestCase):
    def test_same_number_on_an_unbound_parent_is_left_for_generic_handling(self):
        outcome = structured.first_refusal(
            {"opportunity_id": "synthetic", "opportunity_number": "DE-FOA-0003215"},
            b"",
            {},
            detail_fetcher=lambda _identifier: {},
            collector=lambda _data: [],
            download=lambda _url: {},
            extract_containers=extractor,
            as_of="2026-08-20",
        )
        self.assertIsNone(outcome)


if __name__ == "__main__":
    unittest.main()
