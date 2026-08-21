"""Bounded agency-declared child routes accepted at the P9.0 gate.

These are named exceptions over measured official sources, not new generic
parsers. A recognized parent is *claimed*: source failure returns zero children
and never falls through to weaker generic inference. That is the provenance
ladder's first-refusal/fail-closed rule in executable form.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from functools import lru_cache
import hashlib
from io import BytesIO
import re

from scripts import subtopic_records
from scripts.subtopic_segmentation import build_term_map, summarize


ARL_NUMBER = "W911NF-23-S-0001"
GENESIS_NUMBER = "DE-FOA-0003612"
HGEO_NUMBERS = frozenset({
    "DE-FOA-0003627",
    "DE-FOA-0003634",
    "DE-FOA-0003215",
})
HGEO_PARENT_IDS = {
    "DE-FOA-0003627": "363065",
    "DE-FOA-0003634": "363302",
    "DE-FOA-0003215": "363594",
}
ARL_PARENT_ID = "344592"
GENESIS_PARENT_ID = "361526"
ROSES_NUMBER_RE = re.compile(r"^NNH\d{2}ZDA\d{3}[A-Z]", re.IGNORECASE)
ARL_TOPIC_RE = re.compile(r"(?<![A-Za-z0-9])ARL-BAA-\d{4}(?!\d)")

KNOWN_ARL_VERSIONS = {
    "c9ab5dd5a95c0f40f68fa4af8b4600c4534e26a15f09a16662e53fb795ba8b24": 82,
}
KNOWN_GENESIS_VERSIONS = {
    "a2e36829b1c6f1ece1db19e6baf854fb1eff34a41d79efbb6bc60a646a9e3517": (21, 98),
}

HGEO_EXPECTED = {
    "DE-FOA-0003627": (
        ("1a", "Field Test Site Research and Development of Technologies for Enhanced Recovery from Unconventional Oil and Gas Reservoirs"),
        ("1b", "Advanced Characterization of Fracture Propagation, Proppant Behavior, and Well Diagnostics"),
        ("1c", "Enhanced Recovery of Oil and Gas from Unconventional Reservoirs Using Carbon Dioxide (CO2)"),
        ("2", "Advanced Field-Testing of Multi-Scale Produced Water Treatment Technologies & Processes"),
    ),
    "DE-FOA-0003634": (
        ("1", "Enhanced Resource Utilization and Production Technologies (ERUPT)"),
        ("1a", "Laboratory Validation of Catalysts and Unit Operations"),
        ("1b", "Field Validation of Full System Prototypes"),
        ("2", "Resilient Infrastructure Technologies Enhancement (RITE)"),
        ("3", "Hydrocarbon Infrastructure Test Sites (HITS)"),
    ),
    "DE-FOA-0003215": (
        ("1A", "Coal"),
        ("1B", "Oil & Gas"),
        ("1C", "Geothermal"),
    ),
}

HGEO_FILENAMES = {
    "DE-FOA-0003627": re.compile(r"^FundOpp_DE-FOA-0003627_Amd_000003\.pdf$", re.I),
    "DE-FOA-0003634": re.compile(r"^FundOpp_DE-FOA-0003634\.pdf$", re.I),
    "DE-FOA-0003215": re.compile(r"^FundOpp_DE-FOA-0003215\.pdf$", re.I),
}
ARL_FILENAME = re.compile(
    r"^Current Research Topics for DEVCOM ARL BAA .*\.pdf$", re.I
)
GENESIS_FILENAME = re.compile(
    r"^Genesis Mission Phase I Application Template v2\.xlsx$", re.I
)


@dataclass(frozen=True)
class StructuredOutcome:
    claimed: bool
    records: tuple = ()
    document: dict | None = None
    method: str | None = None
    provenance: str | None = None
    reason: str | None = None
    diagnostics: dict = field(default_factory=dict)


@lru_cache(maxsize=1)
def _live_roses_payload():
    from scripts.sources.adapters.nasa_roses import NasaRosesAdapter

    return NasaRosesAdapter().fetch()


def _collapse(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _page_before(text, offset):
    pages = list(re.finditer(r"^=====\s+(\d+)\s+=====$", text[:offset], re.M))
    return int(pages[-1].group(1)) if pages else None


def parse_hgeo(text, opportunity_number):
    """Reproduce the adjudicated 4/5/3 source-specific applicant units."""
    expected = HGEO_EXPECTED.get(opportunity_number)
    if not expected:
        return []
    flat = _collapse(text).casefold()
    children = []
    for ordinal, (code, title) in enumerate(expected, start=1):
        tokens = re.findall(r"[a-z0-9]+", title.casefold())
        # Long source-published titles are the canary. Matching the first six
        # substantive tokens tolerates PDF line wrapping but not structure loss.
        needle = " ".join(tokens[: min(6, len(tokens))])
        if needle not in re.sub(r"[^a-z0-9]+", " ", flat):
            return []
        if opportunity_number == "DE-FOA-0003215":
            marker = rf"Subtopic\s+{re.escape(code)}\s*:\s*{re.escape(title)}"
        else:
            marker = rf"Topic\s+Area\s+{re.escape(code)}\b"
        found = re.search(marker, text, re.I)
        if not found:
            return []
        children.append({
            "code": code,
            "title": title,
            "ordinal": ordinal,
            "summary": (
                f"{title}. The notice identifies this as an applicant-selectable "
                "topic or subtopic."
            ),
            "text": title,
            "page_start": _page_before(text, found.start()),
            "anchor": f"p{_page_before(text, found.start())}" if _page_before(text, found.start()) else None,
        })
    return children


def parse_arl_topics(text):
    """Read the versioned ARL current-topics PDF by its stable announcement IDs."""
    matches = list(ARL_TOPIC_RE.finditer(text or ""))
    if not matches:
        return []
    children = []
    for ordinal, match in enumerate(matches, start=1):
        title_marker = text.rfind("Title:", max(0, match.start() - 1000), match.start())
        if title_marker < 0:
            return []
        line_start = text.rfind("\n", max(0, title_marker - 500), title_marker) + 1
        before = _collapse(text[line_start:title_marker])
        continuation = _collapse(text[title_marker + len("Title:"):match.start()])
        title = _collapse(f"{before} {continuation}")
        if not title or len(title) > 240:
            return []
        next_start = matches[ordinal].start() if ordinal < len(matches) else len(text)
        # The next topic's title precedes its ARL-BAA identifier. Stop at that
        # title rather than letting it leak into this topic's summary.
        if ordinal < len(matches):
            next_title = text.rfind("Title:", match.end(), next_start)
            if next_title >= 0:
                next_start = next_title
        chunk = text[match.end():next_start]
        description = chunk.split("Description:", 1)[1] if "Description:" in chunk else chunk
        children.append({
            "code": match.group(0),
            "title": title,
            "ordinal": ordinal,
            "summary": summarize(description),
            "text": f"{title} {chunk}",
            "page_start": _page_before(text, title_marker),
            "anchor": f"p{_page_before(text, title_marker)}" if _page_before(text, title_marker) else None,
        })
    if len({child["code"] for child in children}) != len(children):
        return []
    return children


def parse_genesis_workbook(content):
    """Genesis-only Phase-I dropdown parser; never a generic XLSX route."""
    from openpyxl import load_workbook

    book = load_workbook(BytesIO(content), read_only=True, data_only=True)
    if "Phase I Summary" not in book.sheetnames or "Focus Areas" not in book.sheetnames:
        return [], {}
    summary_values = " ".join(
        _collapse(value)
        for row in book["Phase I Summary"].iter_rows(values_only=True)
        for value in row
        if value not in (None, "")
    ).casefold()
    if "focus area" not in summary_values or "select from dropdown menu" not in summary_values:
        return [], {}

    focus_re = re.compile(
        r"^(?P<group>\d{1,2})-(?P<letter>[A-Z])\s+"
        r"(?P<challenge>.+?)\s*\|\s*(?P<focus>.+)$"
    )
    rows = []
    for row_number, row in enumerate(
        book["Focus Areas"].iter_rows(values_only=True), start=1
    ):
        for value in row:
            found = focus_re.match(_collapse(value))
            if found:
                rows.append({**found.groupdict(), "row": row_number})
                break
    if not rows:
        return [], {}

    groups = {}
    for row in rows:
        if row["group"] in groups and groups[row["group"]] != row["challenge"]:
            return [], {}
        groups[row["group"]] = row["challenge"]
    children = []
    for group_number in sorted(groups, key=int):
        group_id = f"challenge-{int(group_number)}"
        children.append({
            "code": group_id,
            "code_norm": group_id,
            "title": groups[group_number],
            "summary": f"Genesis Challenge Area {int(group_number)}: {groups[group_number]}.",
            "text": groups[group_number],
            "group_id": group_id,
            "anchor": f"sheet:Focus Areas:group-{int(group_number)}",
        })
    for row in rows:
        group_id = f"challenge-{int(row['group'])}"
        parent_id = subtopic_records.subtopic_id_for("{parent_id}", group_id)
        children.append({
            "code": f"{int(row['group'])}-{row['letter']}",
            "title": row["focus"],
            "summary": f"Challenge Area: {row['challenge']}. Focus Area: {row['focus']}.",
            "text": f"{row['challenge']} {row['focus']}",
            "group_id": group_id,
            # Replaced with the real parent id at the adapter boundary.
            "parent_subtopic_id": parent_id,
            "anchor": f"sheet:Focus Areas:A{row['row']}",
        })
    return children, {"challenge_groups": len(groups), "focus_areas": len(rows)}


def _extract_text(content, document, extract_containers):
    containers, _diagnostics = extract_containers(
        content,
        document.get("content_type"),
        document.get("name"),
        document.get("url"),
    )
    return "\n".join(
        f"===== {container.get('page') or container.get('section') or 'source'} =====\n"
        f"{container.get('text') or ''}"
        for container in containers
    )


def _current_documents(
    record,
    primary_content,
    primary_document,
    *,
    detail_fetcher,
    collector,
    download,
    filename_re=None,
):
    detail = detail_fetcher(str(record.get("opportunity_id")))
    data = detail.get("data", detail) if isinstance(detail, dict) else detail
    attachments = collector(data)
    current_urls = {
        item.get("download_url"): item for item in attachments if item.get("download_url")
    }
    documents = []
    if primary_content and primary_document:
        primary_url = primary_document.get("url")
        meta = current_urls.get(primary_url) or next(
            (
                item for url, item in current_urls.items()
                if url and primary_url and url.rsplit("/", 1)[-1] == primary_url.rsplit("/", 1)[-1]
            ),
            {},
        )
        documents.append((primary_content, {**primary_document, "attachment_id": meta.get("id")}))
    seen = {
        str((primary_document or {}).get("url") or "").rstrip("/").rsplit("/", 1)[-1]
    }
    for item in attachments:
        url = item.get("download_url")
        if not url or (
            filename_re and not filename_re.match(str(item.get("file_name") or ""))
        ):
            continue
        identity = str(url).rstrip("/").rsplit("/", 1)[-1]
        if identity in seen:
            continue
        response = download(url)
        content = response.get("content") or b""
        seen.add(identity)
        documents.append((content, {
            "url": response.get("url") or url,
            "name": item.get("file_name"),
            "content_type": response.get("content_type") or item.get("mime_type"),
            "sha256": hashlib.sha256(content).hexdigest() if content else None,
            "source_kind": "authoritative_notice",
            "attachment_id": item.get("id"),
        }))
    return documents


def _named_document_outcome(
    record,
    primary_content,
    primary_document,
    *,
    filename_re,
    parser,
    provenance,
    method,
    detail_fetcher,
    collector,
    download,
    extract_containers,
    as_of,
):
    try:
        docs = _current_documents(
            record,
            primary_content,
            primary_document,
            detail_fetcher=detail_fetcher,
            collector=collector,
            download=download,
            filename_re=filename_re,
        )
        selected = [(content, document) for content, document in docs
                    if filename_re.match(str(document.get("name") or ""))]
        if len(selected) != 1:
            return StructuredOutcome(True, reason="structured_source_not_unique")
        content, document = selected[0]
        document = {**document, "source_kind": "authoritative_notice"}
        parsed, diagnostics = parser(content, document)
        if not parsed:
            return StructuredOutcome(
                True, document=document, method=method, provenance=provenance,
                reason="structured_source_failed", diagnostics=diagnostics,
            )
        parent_id = str(record.get("opportunity_id"))
        for child in parsed:
            if child.get("parent_subtopic_id"):
                child["parent_subtopic_id"] = child["parent_subtopic_id"].replace(
                    "{parent_id}", parent_id
                )
        built = subtopic_records.build_structured_records(
            record,
            parsed,
            document=document,
            as_of=as_of,
            provenance=provenance,
            confidence="high",
            method=method,
            source_version=document.get("sha256"),
        )
        return StructuredOutcome(
            True, tuple(built), document, method, provenance,
            diagnostics=diagnostics,
        )
    except Exception as exc:  # fail closed for a claimed authoritative route
        return StructuredOutcome(
            True,
            reason=f"structured_error_{type(exc).__name__}",
            method=method,
            provenance=provenance,
        )


def first_refusal(
    record,
    primary_content,
    primary_document,
    *,
    detail_fetcher,
    collector,
    download,
    extract_containers,
    as_of,
    roses_payload=None,
):
    """Return a claimed outcome for a named route, otherwise ``None``."""
    number = str(record.get("opportunity_number") or "")
    if ROSES_NUMBER_RE.match(number):
        from scripts.sources.adapters.nasa_roses import NasaRosesAdapter

        try:
            adapter = NasaRosesAdapter()
            payload = roses_payload() if callable(roses_payload) else roses_payload
            payload = payload or _live_roses_payload()
            rows = adapter.rows(payload)
            health = adapter.check_health(payload, rows)
            if not health["healthy"]:
                return StructuredOutcome(
                    True, method="nasa_roses_table", provenance=subtopic_records.NATIVE,
                    reason="native_source_health_failed", diagnostics=health,
                )
            report = adapter.reconcile(
                rows,
                catalog_records=[record],
                year=payload.get("year"),
                today=date.fromisoformat(as_of),
            )
            if report["review"] or len(report["matched"]) != 1:
                return StructuredOutcome(
                    True, method="nasa_roses_table", provenance=subtopic_records.NATIVE,
                    reason="native_parent_reconciliation_failed",
                    diagnostics={"matched": len(report["matched"]), "review": len(report["review"])},
                )
            element = report["matched"][0]
            raw = payload.get("table3_html") or ""
            document = {
                "url": payload.get("table3_url"),
                "name": f"ROSES-{payload.get('year')} Table 3",
                "sha256": hashlib.sha256(raw.encode("utf-8")).hexdigest(),
                "source_kind": "authoritative_notice",
            }
            built = adapter.subtopic_children(
                rows,
                parent_matches={element["identity"]: record},
                as_of=as_of,
                health=health,
                document=document,
                source_version={
                    "year": payload.get("year"),
                    "amendment": payload.get("amendment"),
                },
            )
            if len(built) != 1:
                return StructuredOutcome(
                    True, method="nasa_roses_table", provenance=subtopic_records.NATIVE,
                    reason="native_child_build_failed", diagnostics=health,
                )
            return StructuredOutcome(
                True, tuple(built), document, "nasa_roses_table",
                subtopic_records.NATIVE, diagnostics=health,
            )
        except Exception as exc:
            return StructuredOutcome(
                True, method="nasa_roses_table", provenance=subtopic_records.NATIVE,
                reason=f"native_error_{type(exc).__name__}",
            )

    parent_id = str(record.get("opportunity_id") or "")
    if number in HGEO_NUMBERS and parent_id == HGEO_PARENT_IDS[number]:
        def parser(content, document):
            text = _extract_text(content, document, extract_containers)
            children = parse_hgeo(text, number)
            return children, {"expected": len(HGEO_EXPECTED[number]), "parsed": len(children)}

        return _named_document_outcome(
            record, primary_content, primary_document,
            filename_re=HGEO_FILENAMES[number], parser=parser,
            provenance=subtopic_records.INLINE, method="hgeo_declared_topics",
            detail_fetcher=detail_fetcher, collector=collector, download=download,
            extract_containers=extract_containers, as_of=as_of,
        )

    if number == ARL_NUMBER and parent_id == ARL_PARENT_ID:
        def parser(content, document):
            text = _extract_text(content, document, extract_containers)
            children = parse_arl_topics(text)
            expected = KNOWN_ARL_VERSIONS.get(document.get("sha256"))
            healthy = len(children) == expected if expected else 40 <= len(children) <= 200
            healthy = healthy and not any("muri" in child["title"].casefold() for child in children)
            return (children if healthy else []), {
                "parsed": len(children), "expected_for_version": expected,
                "muri_topics": sum("muri" in child["title"].casefold() for child in children),
            }

        return _named_document_outcome(
            record, primary_content, primary_document,
            filename_re=ARL_FILENAME, parser=parser,
            provenance=subtopic_records.NATIVE, method="arl_current_topics",
            detail_fetcher=detail_fetcher, collector=collector, download=download,
            extract_containers=extract_containers, as_of=as_of,
        )

    if number == GENESIS_NUMBER and parent_id == GENESIS_PARENT_ID:
        def parser(content, document):
            children, diagnostics = parse_genesis_workbook(content)
            expected = KNOWN_GENESIS_VERSIONS.get(document.get("sha256"))
            observed = (diagnostics.get("challenge_groups"), diagnostics.get("focus_areas"))
            healthy = observed == expected if expected else (
                observed[0] is not None and 1 <= observed[0] <= 50
                and observed[1] is not None and 50 <= observed[1] <= 200
            )
            diagnostics["expected_for_version"] = expected
            return (children if healthy else []), diagnostics

        return _named_document_outcome(
            record, primary_content, primary_document,
            filename_re=GENESIS_FILENAME, parser=parser,
            provenance=subtopic_records.NATIVE, method="genesis_focus_workbook",
            detail_fetcher=detail_fetcher, collector=collector, download=download,
            extract_containers=extract_containers, as_of=as_of,
        )

    return None
