"""Build the compact Team Match faculty directory from the reviewed workbook.

The workbook is an import artifact, not a browser dependency.  The generated
JavaScript keeps only the searchable faculty projection and its auditable
controlled-term mappings.  Existing curated Chemical & Sustainability
Engineering profiles are joined by identity and remain the matching authority.
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import dataclass
import gzip
import hashlib
import json
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook


EXPECTED_SOURCE_SHA256 = "4cc24fad355c5716a462b93e1f60d0c7d55d9368d7cfede330ff41daa36af130"
EXPECTED_COUNTS = {
    "source_profiles": 156,
    "retained_profiles": 1,
    "searchable_profiles": 157,
    "controlled_terms": 202,
    "primary_mappings": 460,
    "supporting_mappings": 94,
    "matching_available": 157,
    "curated_profiles": 13,
}
DEFAULT_OUTPUT = Path("data/hajim-faculty-directory.js")
DEFAULT_HTML = Path("team_match.html")
RAW_SIZE_BUDGET = 275_000
GZIP_SIZE_BUDGET = 48_000
ASSIGNMENT_PREFIX = "globalThis.HAJIM_FACULTY_DIRECTORY="

CURATED_PROFILE_KEYS = (
    "Mitchell Anthamatten",
    "Yasemin Basdogan",
    "Pooja Rajendra Bhalode",
    "Siddharth Deshpande",
    "Gang Fan",
    "David G. Foster",
    "Darren Lipomi",
    "Allison J. Lopatkin",
    "Astrid M. Muller",
    "Marc D. Porosoff",
    "Alexander A. Shestopalov",
    "Wyatt E. Tenhaeff",
    "Matthew Z. Yates",
)

FACULTY_HEADERS = (
    "Faculty Name",
    "Primary / Home Unit",
    "Faculty Relationship",
    "Academic Rank / Appointments",
    "Hajim Faculty Roster(s)",
    "Validated Research Profile",
    "Precision Match Terms",
    "Email",
    "Lab / Faculty Website",
    "Validation Evidence URL(s)",
    "Checked Date",
    "Curated Controlled Match Terms",
)
VOCABULARY_HEADERS = (
    "Term ID",
    "Category",
    "Controlled Match Term",
    "Term Type",
    "Matching Role",
    "Search Aliases",
    "Definition / Scope",
    "Faculty Count",
    "Use Rule",
)
MAPPING_HEADERS = (
    "Faculty Name",
    "Primary / Home Unit",
    "Term ID",
    "Controlled Match Term",
    "Category",
    "Term Type",
    "Matching Role",
    "Mapping Confidence",
    "Matched Phrase",
    "Mapping Evidence",
    "Profile Source",
    "Graph Eligible",
)
RETAINED_HEADERS = (
    "Faculty Name",
    "Primary / Home Unit",
    "Relationship",
    "Curated Research Profile",
    "Curated Precision Terms",
    "Curated Controlled Match Terms",
)


class DirectoryContractError(ValueError):
    """Raised when the import or committed projection violates its contract."""


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def split_terms(value: object) -> list[str]:
    return [item for item in (clean(part) for part in clean(value).split(";")) if item]


def identity_text(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", clean(value))
    return re.sub(r"[^a-z0-9]+", "", normalized.encode("ascii", "ignore").decode().lower())


def faculty_id(name: str) -> str:
    normalized = unicodedata.normalize("NFKD", name)
    slug = re.sub(
        r"[^a-z0-9]+",
        "-",
        normalized.encode("ascii", "ignore").decode().lower(),
    ).strip("-")
    if not slug:
        raise DirectoryContractError(f"Could not derive an identity for {name!r}.")
    return f"hajim:{slug}"


def canonical_bytes(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def source_sha256(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def sheet_rows(workbook, name: str, expected_headers: tuple[str, ...]) -> list[dict[str, object]]:
    if name not in workbook.sheetnames:
        raise DirectoryContractError(f"Workbook is missing the {name!r} sheet.")
    values = workbook[name].iter_rows(values_only=True)
    try:
        actual_headers = tuple(clean(value) for value in next(values))
    except StopIteration as error:
        raise DirectoryContractError(f"Workbook sheet {name!r} is empty.") from error
    if actual_headers != expected_headers:
        raise DirectoryContractError(
            f"Workbook sheet {name!r} headers changed: {actual_headers!r}."
        )
    rows = []
    for values_row in values:
        if not any(clean(value) for value in values_row):
            continue
        rows.append(dict(zip(expected_headers, values_row, strict=True)))
    return rows


def curated_key_for(name: str) -> str:
    lookup = {identity_text(item): item for item in CURATED_PROFILE_KEYS}
    return lookup.get(identity_text(name), "")


@dataclass(frozen=True)
class WorkbookData:
    faculty: list[dict[str, object]]
    vocabulary: list[dict[str, object]]
    mappings: list[dict[str, object]]
    retained: list[dict[str, object]]


def read_workbook(path: Path) -> WorkbookData:
    actual_hash = source_sha256(path)
    if actual_hash != EXPECTED_SOURCE_SHA256:
        raise DirectoryContractError(
            "Workbook SHA-256 does not match the reviewed source: "
            f"expected {EXPECTED_SOURCE_SHA256}, got {actual_hash}."
        )
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return WorkbookData(
            faculty=sheet_rows(workbook, "Faculty Profiles", FACULTY_HEADERS),
            vocabulary=sheet_rows(workbook, "Controlled Vocabulary", VOCABULARY_HEADERS),
            mappings=sheet_rows(workbook, "Faculty-Term Map", MAPPING_HEADERS),
            retained=sheet_rows(workbook, "Retained Curated Profiles", RETAINED_HEADERS),
        )
    finally:
        workbook.close()


def mapping_projection(row: dict[str, object]) -> dict[str, str]:
    return {
        "term_id": clean(row["Term ID"]),
        "source_phrase": clean(row["Matched Phrase"]),
        "evidence": clean(row["Mapping Evidence"]),
        "profile_source": clean(row["Profile Source"]),
    }


def build_projection(data: WorkbookData) -> dict[str, object]:
    faculty_names = [clean(row["Faculty Name"]) for row in data.faculty]
    retained_names = [clean(row["Faculty Name"]) for row in data.retained]
    searchable_names = faculty_names + retained_names
    if any(not name for name in searchable_names):
        raise DirectoryContractError("Every faculty profile must have a name.")
    duplicate_names = [name for name, count in Counter(searchable_names).items() if count != 1]
    if duplicate_names:
        raise DirectoryContractError(f"Faculty identities are duplicated: {duplicate_names!r}.")
    if any(identity_text(name) == "melodieilawton" for name in searchable_names):
        raise DirectoryContractError("Melodie I. Lawton must not be a faculty profile.")

    term_rows: dict[str, dict[str, object]] = {}
    for row in data.vocabulary:
        term_id = clean(row["Term ID"])
        if not re.fullmatch(r"CV\d{3}", term_id) or term_id in term_rows:
            raise DirectoryContractError(f"Controlled term identity is invalid or duplicated: {term_id!r}.")
        role = clean(row["Matching Role"]).lower()
        if role not in {"primary anchor", "supporting context"}:
            raise DirectoryContractError(f"Controlled term {term_id} has an invalid matching role.")
        term_rows[term_id] = {
            "id": term_id,
            "label": clean(row["Controlled Match Term"]),
            "category": clean(row["Category"]),
            "type": clean(row["Term Type"]),
            "role": role.replace(" ", "_"),
            "aliases": split_terms(row["Search Aliases"]),
            "scope": clean(row["Definition / Scope"]),
        }

    mappings_by_name: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(
        lambda: {"primary": [], "context": []}
    )
    seen_mappings: set[tuple[str, str]] = set()
    for row in data.mappings:
        name = clean(row["Faculty Name"])
        term_id = clean(row["Term ID"])
        if name not in searchable_names:
            raise DirectoryContractError(f"Mapping references unknown faculty profile {name!r}.")
        term = term_rows.get(term_id)
        if not term:
            raise DirectoryContractError(f"Mapping references unknown controlled term {term_id!r}.")
        key = (name, term_id)
        if key in seen_mappings:
            raise DirectoryContractError(f"Duplicate faculty-term mapping: {key!r}.")
        seen_mappings.add(key)
        row_role = clean(row["Matching Role"]).lower()
        graph_eligible = clean(row["Graph Eligible"]).lower()
        if row_role == "primary anchor" and graph_eligible == "yes":
            bucket = "primary"
        elif row_role == "supporting context" and graph_eligible.startswith("no"):
            bucket = "context"
        else:
            raise DirectoryContractError(
                f"Mapping {name!r}/{term_id} has an inconsistent admission role."
            )
        if term["role"] != row_role.replace(" ", "_"):
            raise DirectoryContractError(f"Mapping {name!r}/{term_id} conflicts with the vocabulary role.")
        if clean(row["Controlled Match Term"]) != term["label"]:
            raise DirectoryContractError(f"Mapping {name!r}/{term_id} changed its controlled label.")
        projection = mapping_projection(row)
        if not projection["source_phrase"] or not projection["evidence"]:
            raise DirectoryContractError(f"Mapping {name!r}/{term_id} lacks auditable evidence.")
        mappings_by_name[name][bucket].append(projection)

    profiles = []
    ids: set[str] = set()
    profile_rows = [
        {
            "name": clean(row["Faculty Name"]),
            "unit": clean(row["Primary / Home Unit"]),
            "relationship": clean(row["Faculty Relationship"]),
            "appointments": clean(row["Academic Rank / Appointments"]),
            "rosters": split_terms(row["Hajim Faculty Roster(s)"]),
            "summary": clean(row["Validated Research Profile"]),
        }
        for row in data.faculty
    ] + [
        {
            "name": clean(row["Faculty Name"]),
            "unit": clean(row["Primary / Home Unit"]),
            "relationship": clean(row["Relationship"]),
            "appointments": "",
            "rosters": [clean(row["Primary / Home Unit"])],
            "summary": clean(row["Curated Research Profile"]),
        }
        for row in data.retained
    ]
    for row in sorted(profile_rows, key=lambda item: (identity_text(item["name"]), item["name"])):
        identifier = faculty_id(row["name"])
        if identifier in ids:
            raise DirectoryContractError(f"Stable faculty identity is duplicated: {identifier!r}.")
        ids.add(identifier)
        mapped = mappings_by_name[row["name"]]
        mapped["primary"].sort(key=lambda item: item["term_id"])
        mapped["context"].sort(key=lambda item: item["term_id"])
        profiles.append({
            "id": identifier,
            **row,
            "curated_profile_key": curated_key_for(row["name"]),
            "matching_available": bool(mapped["primary"]),
            "primary": mapped["primary"],
            "context": mapped["context"],
        })

    counts = {
        "source_profiles": len(data.faculty),
        "retained_profiles": len(data.retained),
        "searchable_profiles": len(profiles),
        "controlled_terms": len(term_rows),
        "primary_mappings": sum(len(item["primary"]) for item in profiles),
        "supporting_mappings": sum(len(item["context"]) for item in profiles),
        "matching_available": sum(bool(item["primary"]) for item in profiles),
        "curated_profiles": sum(bool(item["curated_profile_key"]) for item in profiles),
    }
    if counts != EXPECTED_COUNTS:
        raise DirectoryContractError(
            f"Workbook counts changed: expected {EXPECTED_COUNTS!r}, got {counts!r}."
        )

    core: dict[str, object] = {
        "schema_version": 1,
        "source_sha256": EXPECTED_SOURCE_SHA256,
        "source_checked_date": "2026-08-29",
        "counts": counts,
        "terms": [term_rows[key] for key in sorted(term_rows)],
        "profiles": profiles,
    }
    return {**core, "generation_identity": sha256_bytes(canonical_bytes(core))}


def asset_bytes(payload: dict[str, object]) -> bytes:
    return (
        "/* Generated by scripts/hajim_faculty_directory.py. Do not edit by hand. */\n"
        f"{ASSIGNMENT_PREFIX}{canonical_bytes(payload).decode('utf-8')};\n"
    ).encode("utf-8")


def parse_asset(value: bytes) -> dict[str, object]:
    text = value.decode("utf-8")
    start = text.find(ASSIGNMENT_PREFIX)
    if start < 0 or not text.rstrip().endswith(";"):
        raise DirectoryContractError("Generated directory is not the expected JavaScript assignment.")
    serialized = text[start + len(ASSIGNMENT_PREFIX):].strip()
    if serialized.endswith(";"):
        serialized = serialized[:-1]
    try:
        payload = json.loads(serialized)
    except json.JSONDecodeError as error:
        raise DirectoryContractError("Generated directory contains invalid JSON.") from error
    if not isinstance(payload, dict):
        raise DirectoryContractError("Generated directory payload must be an object.")
    return payload


def validate_payload(payload: dict[str, object]) -> dict[str, int | str]:
    identity = clean(payload.get("generation_identity"))
    core = {key: value for key, value in payload.items() if key != "generation_identity"}
    actual_identity = sha256_bytes(canonical_bytes(core))
    if identity != actual_identity:
        raise DirectoryContractError(
            f"Generated identity mismatch: expected {actual_identity}, got {identity}."
        )
    if payload.get("schema_version") != 1:
        raise DirectoryContractError("Generated directory schema is incompatible.")
    if payload.get("source_sha256") != EXPECTED_SOURCE_SHA256:
        raise DirectoryContractError("Generated directory points to an unreviewed workbook.")
    if payload.get("counts") != EXPECTED_COUNTS:
        raise DirectoryContractError("Generated directory counts changed.")
    profiles = payload.get("profiles")
    terms = payload.get("terms")
    if not isinstance(profiles, list) or not isinstance(terms, list):
        raise DirectoryContractError("Generated directory collections are missing.")
    if len({clean(item.get("id")) for item in profiles if isinstance(item, dict)}) != len(profiles):
        raise DirectoryContractError("Generated directory has duplicate faculty identities.")
    if any("melodie" in canonical_bytes(item).decode("utf-8").lower() for item in profiles):
        raise DirectoryContractError("Generated directory must not contain Melodie Lawton.")
    curated = {clean(item.get("curated_profile_key")) for item in profiles if isinstance(item, dict)}
    if curated - {"", *CURATED_PROFILE_KEYS} or curated - {""} != set(CURATED_PROFILE_KEYS):
        raise DirectoryContractError("Generated directory does not join exactly the curated ChemE identities.")
    return {
        **EXPECTED_COUNTS,
        "generation_identity": identity,
        "source_sha256": EXPECTED_SOURCE_SHA256,
    }


def size_report(value: bytes) -> tuple[int, int]:
    raw = len(value)
    compressed = len(gzip.compress(value, mtime=0))
    if raw > RAW_SIZE_BUDGET or compressed > GZIP_SIZE_BUDGET:
        raise DirectoryContractError(
            f"Directory exceeds its budget: {raw} raw/{compressed} gzip bytes; "
            f"limits are {RAW_SIZE_BUDGET}/{GZIP_SIZE_BUDGET}."
        )
    return raw, compressed


def synchronize_html_generation(path: Path, identity: str, *, write: bool) -> None:
    text = path.read_text(encoding="utf-8")
    pattern = re.compile(
        r'(<meta name="hajim-faculty-directory-generation" content=")[a-f0-9]{64}("\s*/>)'
    )
    if len(pattern.findall(text)) != 1:
        raise DirectoryContractError("Team Match must declare exactly one faculty-directory generation.")
    expected = pattern.sub(rf"\g<1>{identity}\g<2>", text)
    if write:
        if expected != text:
            path.write_text(expected, encoding="utf-8")
    elif expected != text:
        raise DirectoryContractError("Team Match points to a different faculty-directory generation.")


def run(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    action = parser.add_mutually_exclusive_group(required=True)
    action.add_argument("--write", action="store_true")
    action.add_argument("--check", action="store_true")
    parser.add_argument("--source", type=Path)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--html", type=Path, default=DEFAULT_HTML)
    args = parser.parse_args(argv)

    if args.write:
        if not args.source:
            parser.error("--write requires --source")
        payload = build_projection(read_workbook(args.source))
        value = asset_bytes(payload)
        size_report(value)
        args.out.parent.mkdir(parents=True, exist_ok=True)
        args.out.write_bytes(value)
        synchronize_html_generation(args.html, payload["generation_identity"], write=True)
        status = "written"
    else:
        if args.source:
            parser.error("--check validates the committed projection and does not accept --source")
        value = args.out.read_bytes()
        payload = parse_asset(value)
        status = "verified"

    summary = validate_payload(payload)
    synchronize_html_generation(args.html, payload["generation_identity"], write=False)
    raw, compressed = size_report(value)
    print(json.dumps({"status": status, **summary, "raw_bytes": raw, "gzip_bytes": compressed}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
