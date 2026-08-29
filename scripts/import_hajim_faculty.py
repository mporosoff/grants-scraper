"""Deterministically import the reviewed Hajim faculty workbook.

The workbook is a handoff artifact, not a runtime dependency.  This module
validates the reviewed snapshot and emits the canonical repository JSON used
by the faculty-match generator.
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime
import hashlib
import json
from pathlib import Path
import re
import unicodedata
from urllib.parse import urlsplit

from openpyxl import load_workbook


SHEET_NAME = "Faculty Profiles"
HEADERS = (
    "Faculty Name",
    "Primary / Home Unit",
    "Faculty Relationship",
    "Academic Rank / Appointments",
    "Hajim Faculty Roster(s)",
    "Research Interests (website text, lightly normalized)",
    "Derived Research Theme(s)",
    "Email",
    "Lab / Faculty Website",
    "Source Faculty Page URL(s)",
    "Checked Date",
)
RELATIONSHIPS = {
    "Hajim primary/core faculty": "hajim_primary_core",
    "Hajim research faculty": "hajim_research",
    "Joint Hajim appointment / program faculty": "joint_hajim_or_program",
    "Materials Science program faculty (non-Hajim home)": "materials_science_non_hajim_home",
}
MISSING_INTEREST_SENTINEL = "Not listed on source faculty page"
EXPECTED_COUNTS = {
    "hajim_primary_core": 115,
    "hajim_research": 11,
    "joint_hajim_or_program": 19,
    "materials_science_non_hajim_home": 11,
}
EXPECTED_RECORDS = 156
EXPECTED_RANKABLE = 145
EXPECTED_CHECKED_DATE = "2026-08-28"
EXPECTED_SOURCE_SHA256 = "f625ec89beabcfe7a7c178b83dcd9ca6737be455fc70c3b00f06882f2d6114fc"
SCHEMA_VERSION = 1


class FacultyImportError(ValueError):
    """Raised when a workbook or canonical payload violates the contract."""


def _text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", str(value))).strip()


def _checked_date(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return _text(value)


def _split(value: object, delimiter: str = ";") -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [_text(item) for item in text.split(delimiter) if _text(item)]


def _slug(name: str) -> str:
    folded = unicodedata.normalize("NFKD", name)
    ascii_name = "".join(char for char in folded if not unicodedata.combining(char))
    value = re.sub(r"[^a-z0-9]+", "-", ascii_name.lower()).strip("-")
    if not value:
        raise FacultyImportError(f"Cannot generate faculty_id for {name!r}")
    return value


def _is_https(value: str) -> bool:
    parsed = urlsplit(value)
    return parsed.scheme.lower() == "https" and bool(parsed.netloc)


def _stable_id(name: str, email: str, occupied: set[str]) -> str:
    base = _slug(name)
    if base not in occupied:
        occupied.add(base)
        return base
    suffix = hashlib.sha256(email.encode("utf-8")).hexdigest()[:8]
    candidate = f"{base}-{suffix}"
    counter = 2
    while candidate in occupied:
        candidate = f"{base}-{suffix}-{counter}"
        counter += 1
    occupied.add(candidate)
    return candidate


def validate_payload(payload: dict, *, require_snapshot: bool = False) -> None:
    if not isinstance(payload, dict) or payload.get("schema_version") != SCHEMA_VERSION:
        raise FacultyImportError("Canonical faculty schema_version must be 1")
    source = payload.get("source") or {}
    profiles = payload.get("profiles")
    if not isinstance(profiles, list):
        raise FacultyImportError("Canonical faculty profiles must be a list")
    seen_ids: set[str] = set()
    seen_emails: set[str] = set()
    relationship_counts: Counter[str] = Counter()
    unrankable = 0
    previous_sort_key: tuple[str, str] | None = None
    for profile in profiles:
        required = ("faculty_id", "name", "home_unit", "relationship", "relationship_label",
                    "appointment_text", "appointments", "rosters", "research_interests_text",
                    "research_phrases", "derived_themes", "email", "source_urls", "checked_date",
                    "rankable")
        missing = [key for key in required if key not in profile]
        if missing:
            raise FacultyImportError(f"Profile is missing fields: {', '.join(missing)}")
        faculty_id = profile["faculty_id"]
        email = str(profile["email"]).lower()
        if not profile["name"] or not profile["home_unit"] or not email:
            raise FacultyImportError("Name, home unit, and email are required")
        if faculty_id in seen_ids or email in seen_emails:
            raise FacultyImportError(f"Duplicate faculty identity: {faculty_id} / {email}")
        seen_ids.add(faculty_id)
        seen_emails.add(email)
        relationship = profile["relationship"]
        if relationship not in EXPECTED_COUNTS:
            raise FacultyImportError(f"Unknown relationship enum: {relationship!r}")
        relationship_counts[relationship] += 1
        urls = list(profile["source_urls"])
        if not urls or any(not _is_https(url) for url in urls):
            raise FacultyImportError(f"Profile {faculty_id} must have only HTTPS source URLs")
        website = profile.get("website_url")
        if website is not None and not _is_https(website):
            raise FacultyImportError(f"Profile {faculty_id} has a non-HTTPS website URL")
        rankable = bool(profile["rankable"])
        if rankable != bool(profile["research_phrases"]):
            raise FacultyImportError(f"Profile {faculty_id} rankable state conflicts with interests")
        if not rankable:
            unrankable += 1
            if profile["research_interests_text"] != MISSING_INTEREST_SENTINEL:
                raise FacultyImportError(f"Profile {faculty_id} lacks the exact missing-interest sentinel")
        sort_key = (unicodedata.normalize("NFC", profile["name"]).casefold(), faculty_id)
        if previous_sort_key is not None and sort_key < previous_sort_key:
            raise FacultyImportError("Profiles are not in deterministic name order")
        previous_sort_key = sort_key
    if source.get("record_count") != len(profiles):
        raise FacultyImportError("Source record_count does not match profiles")
    rankable_count = len(profiles) - unrankable
    if source.get("rankable_record_count") != rankable_count or source.get("unlisted_interest_count") != unrankable:
        raise FacultyImportError("Source rankable/unlisted counts do not match profiles")
    normalized_counts = {key: relationship_counts[key] for key in EXPECTED_COUNTS}
    if payload.get("counts") != normalized_counts:
        raise FacultyImportError("Relationship counts do not match profiles")
    if require_snapshot:
        if len(profiles) != EXPECTED_RECORDS or rankable_count != EXPECTED_RANKABLE or unrankable != 11:
            raise FacultyImportError("Reviewed snapshot must contain 156 total / 145 rankable / 11 unrankable profiles")
        if dict(relationship_counts) != EXPECTED_COUNTS:
            raise FacultyImportError(f"Reviewed snapshot relationship counts must be {EXPECTED_COUNTS}")
        if source.get("checked_date") != EXPECTED_CHECKED_DATE:
            raise FacultyImportError("Reviewed snapshot checked date is incompatible")
        if source.get("sha256") != EXPECTED_SOURCE_SHA256:
            raise FacultyImportError("Reviewed snapshot workbook SHA-256 is incompatible")
        if any(profile["checked_date"] != EXPECTED_CHECKED_DATE for profile in profiles):
            raise FacultyImportError("Every reviewed profile must use the reviewed checked date")


def import_workbook(path: str | Path, *, require_snapshot: bool = True) -> dict:
    workbook_path = Path(path)
    source_bytes = workbook_path.read_bytes()
    source_sha = hashlib.sha256(source_bytes).hexdigest()
    # Normal mode fully reads the small reviewed workbook and does not retain a
    # Windows file handle after parsing, which keeps importer tests portable.
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise FacultyImportError(f"Workbook must contain the exact {SHEET_NAME!r} sheet")
    sheet = workbook[SHEET_NAME]
    rows = sheet.iter_rows(values_only=True)
    headers = tuple(_text(value) for value in next(rows, ()))
    if headers != HEADERS:
        raise FacultyImportError(f"Unexpected {SHEET_NAME} headers: {headers!r}")

    pending: list[dict] = []
    emails: set[str] = set()
    for row_number, values in enumerate(rows, start=2):
        if not any(_text(value) for value in values):
            continue
        record = dict(zip(HEADERS, values, strict=True))
        name = _text(record[HEADERS[0]])
        home_unit = _text(record[HEADERS[1]])
        relationship_label = _text(record[HEADERS[2]])
        relationship = RELATIONSHIPS.get(relationship_label)
        if relationship is None:
            raise FacultyImportError(f"Row {row_number} has unknown relationship {relationship_label!r}")
        appointment_text = _text(record[HEADERS[3]])
        rosters = _split(record[HEADERS[4]])
        interests = _text(record[HEADERS[5]])
        rankable = interests != MISSING_INTEREST_SENTINEL
        phrases = _split(interests) if rankable else []
        themes = _split(record[HEADERS[6]])
        email = _text(record[HEADERS[7]]).lower()
        website = _text(record[HEADERS[8]]) or None
        source_urls = _split(record[HEADERS[9]], " | ")
        checked = _checked_date(record[HEADERS[10]])
        if not name or not home_unit or not email or not source_urls or not checked:
            raise FacultyImportError(f"Row {row_number} is missing a required field")
        if email in emails:
            raise FacultyImportError(f"Row {row_number} duplicates email {email}")
        emails.add(email)
        if website is not None and not _is_https(website):
            raise FacultyImportError(f"Row {row_number} has a non-HTTPS website URL")
        if any(not _is_https(url) for url in source_urls):
            raise FacultyImportError(f"Row {row_number} has a non-HTTPS source URL")
        pending.append({
            "name": name,
            "home_unit": home_unit,
            "relationship": relationship,
            "relationship_label": relationship_label,
            "appointment_text": appointment_text,
            "appointments": _split(appointment_text),
            "rosters": sorted(set(rosters), key=str.casefold),
            "research_interests_text": interests,
            "research_phrases": phrases,
            "derived_themes": sorted(set(themes), key=str.casefold),
            "email": email,
            "website_url": website,
            "source_urls": sorted(set(source_urls)),
            "checked_date": checked,
            "rankable": rankable,
        })

    pending.sort(key=lambda item: (item["name"].casefold(), item["email"]))
    occupied: set[str] = set()
    profiles = []
    for item in pending:
        profiles.append({"faculty_id": _stable_id(item["name"], item["email"], occupied), **item})
    counts = Counter(profile["relationship"] for profile in profiles)
    rankable_count = sum(1 for profile in profiles if profile["rankable"])
    checked_dates = sorted({profile["checked_date"] for profile in profiles})
    payload = {
        "schema_version": SCHEMA_VERSION,
        "source": {
            "kind": "reviewed_hajim_faculty_xlsx",
            "filename": workbook_path.name,
            "sha256": source_sha,
            "checked_date": checked_dates[0] if len(checked_dates) == 1 else "",
            "record_count": len(profiles),
            "rankable_record_count": rankable_count,
            "unlisted_interest_count": len(profiles) - rankable_count,
        },
        "counts": {key: counts[key] for key in EXPECTED_COUNTS},
        "profiles": profiles,
    }
    validate_payload(payload, require_snapshot=require_snapshot)
    return payload


def canonical_bytes(payload: dict) -> bytes:
    return (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")


def write_import(path: str | Path, output: str | Path, *, require_snapshot: bool = True) -> bytes:
    data = canonical_bytes(import_workbook(path, require_snapshot=require_snapshot))
    Path(output).write_bytes(data)
    return data


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Reviewed Hajim faculty .xlsx artifact")
    parser.add_argument("--out", default="config/hajim_faculty.json")
    parser.add_argument("--allow-non-snapshot", action="store_true", help="Use only for compact unit-test fixtures")
    args = parser.parse_args(argv)
    data = write_import(args.workbook, args.out, require_snapshot=not args.allow_non_snapshot)
    print(f"Wrote {args.out} ({len(data)} bytes)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
