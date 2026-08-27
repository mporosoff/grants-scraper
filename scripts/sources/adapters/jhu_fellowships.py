"""Johns Hopkins RDT fellowship repositories (graduate / postdoc / early-career).

JHU's Research Development & Translation team maintains three continuously
updated spreadsheets of student and early-career funding, organized by audience.
Each sub-page links to an .xlsx (currently via a bit.ly). We ride on their
curation to fill the fellowship buckets our federal Grants.gov catalog doesn't
cover (GRFP, NDSEG, Hertz, Ford, foundation fellowships, ...).

Staying current: the adapter scrapes each sub-page for its current spreadsheet
link every run and downloads it fresh, so it tracks JHU's updates automatically
rather than a one-time static import.

Avoiding duplicates / mess:
- Each row is audience-tagged via ``applicant_types`` (Graduate students /
  Postdoctoral researchers / Early-career faculty), so items land in the
  fellowship buckets rather than flooding the default view.
- The same sponsor/program appearing in multiple workbooks is merged into one
  record with all applicable audiences.
- On the *early-career faculty* sheet (the one that overlaps our federal catalog
  most — CAREER, NIH ESI, etc.) federal sponsors are dropped, keeping only the
  foundation/private awards we don't already have.
- The merge layer still lets Grants.gov win and drops same-id repeats.

The production adapter retries transient failures, falls back to the official
short links published by JHU, and validates that all three audience workbooks
still contain their expected raw row volume. Only rows with an exact current or
future deadline, or an explicit rolling deadline, are published. A valid
current result may therefore be empty. If JHU blocks an automated refresh, the
catalog publishes zero JHU records rather than reviving an unverifiable snapshot.
"""

from __future__ import annotations

import datetime as _dt
from collections import Counter
import hashlib
import http.cookiejar
import io
import re
import time
import urllib.error
import urllib.request
from typing import Iterable, Optional

from ..base import CanonicalOpportunity, SourceAdapter
from ..http import USER_AGENT
from ..registry import register

SHEETS = [
    {"audience": "grad",
     "page": "https://research.jhu.edu/rdt/funding-opportunities/graduate/",
     "direct_sheet": "https://research.jhu.edu/wp-content/uploads/2026/07/GradFundingOpps7126.xlsx",
     "fallback_sheet": "https://bit.ly/GradFundingOpps7126",
     "snapshot_date": "2026-07-01",
     "applicant_types": ["Graduate students"], "drop_federal": False},
    {"audience": "postdoc",
     "page": "https://research.jhu.edu/rdt/funding-opportunities/postdoctoral/",
     "direct_sheet": "https://research.jhu.edu/wp-content/uploads/2026/07/PostdocFundingOpps7126.xlsx",
     "fallback_sheet": "https://bit.ly/PostdocFundingOpps7126",
     "snapshot_date": "2026-07-01",
     "applicant_types": ["Postdoctoral researchers"], "drop_federal": False},
    {"audience": "faculty",
     "page": "https://research.jhu.edu/rdt/funding-opportunities/early-career/",
     "direct_sheet": "https://research.jhu.edu/wp-content/uploads/2026/07/ECFopps7126.xlsx",
     "fallback_sheet": "https://bit.ly/ECFopps7126",
     "snapshot_date": "2026-07-01",
     "applicant_types": ["Early-career faculty"], "drop_federal": True},
]

# The page-published 7/1/26 workbooks are a bounded compatibility snapshot
# while JHU's category pages are under construction. Two calendar months is
# long enough to bridge the observed outage without presenting one fixed file
# as a perpetually fresh source. A newer page-discovered workbook is not bound
# by this constant.
PINNED_WORKBOOK_MAX_AGE_DAYS = 62

MIN_ROWS_PER_AUDIENCE = 50
FETCH_ATTEMPTS = 3

_FEDERAL_RE = re.compile(
    r"national science foundation|\bnsf\b|national institutes of health|\bnih\b"
    r"|department of energy|\bdoe\b|department of defense|\bdod\b|\bdarpa\b"
    r"|\bnasa\b|department of agriculture|\busda\b|\bepa\b|\bnist\b|\bnoaa\b"
    r"|office of naval research|air force|army research|\bnih\b|\bneh\b|\bnea\b"
    r"|department of education|\bnnsa\b|\barpa",
    re.IGNORECASE,
)

_MONEY_RE = re.compile(r"\$\s?([\d][\d,]{2,})(?:\s?(million|m|k))?", re.IGNORECASE)
_MONTH_RE = (
    r"(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|"
    r"Dec(?:ember)?)"
)
_DATE_RE = re.compile(
    rf"({_MONTH_RE}\s+\d{{1,2}},?\s+\d{{4}}|\d{{4}}-\d{{2}}-\d{{2}}|"
    rf"\d{{1,2}}/\d{{1,2}}/(?:\d{{4}}|\d{{2}}))",
    re.IGNORECASE,
)
_ROLLING_RE = re.compile(
    r"\b(?:rolling|ongoing|continuous(?:ly)?|any\s+time|year[- ]round)\b",
    re.IGNORECASE,
)
_LINK_RE = re.compile(r"""https?://[^\s"'<>]+""", re.IGNORECASE)


def _clean_url(target: Optional[str], fallback: str) -> str:
    """Salvage a real http(s) URL from a cell hyperlink; JHU sheets sometimes
    carry mangled Outlook-cache paths with an embedded URL."""
    if not target:
        return fallback
    text = str(target)
    idx = text.lower().find("http")
    if idx == -1:
        return fallback
    url = text[idx:].strip()
    url = re.sub(r"^(https?):/(?!/)", r"\1://", url)   # fix "http:/x" -> "http://x"
    return url if "." in url else fallback


def _money(value) -> Optional[str]:
    if not value:
        return None
    m = _MONEY_RE.search(str(value))
    if not m:
        return None
    try:
        n = float(m.group(1).replace(",", ""))
    except ValueError:
        return None
    scale = (m.group(2) or "").lower()
    if scale in ("million", "m"):
        n *= 1_000_000
    elif scale == "k":
        n *= 1_000
    return str(int(n))


def _parse_date_token(token: str) -> Optional[_dt.date]:
    cleaned = re.sub(r"\s+", " ", str(token or "")).strip()
    for fmt in (
        "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y",
        "%B %d, %Y", "%B %d %Y", "%b %d, %Y", "%b %d %Y",
    ):
        try:
            return _dt.datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    return None


def _deadline_metadata(value, as_of: _dt.date) -> tuple[Optional[str], Optional[str], str]:
    """Return ``(close_date, note, status)`` without inventing recurrence.

    Exact past dates are expired. Exact current/future dates are publishable.
    Explicitly rolling rows remain current only while a fresh workbook still
    contains them. Undated/TBD/ambiguous rows are excluded because they cannot
    expire automatically or be verified as open.
    """
    if value is None or value == "":
        return None, None, "unverified"
    if isinstance(value, _dt.datetime):
        value = value.date()
    if isinstance(value, _dt.date):
        note = "Deadline " + value.strftime("%b %d, %Y").replace(" 0", " ")
        return value.isoformat(), note, "current" if value >= as_of else "expired"
    text = re.sub(r"\s+", " ", str(value)).strip()
    note = text[:300] or None
    parsed_dates = [
        parsed
        for parsed in (_parse_date_token(match.group(1)) for match in _DATE_RE.finditer(text))
        if parsed is not None
    ]
    if parsed_dates:
        deadline = max(parsed_dates)
        return (
            deadline.isoformat(), note,
            "current" if deadline >= as_of else "expired",
        )
    if _ROLLING_RE.search(text):
        return None, note, "rolling"
    return None, note, "unverified"


def _colmap(header_cells: list) -> dict:
    """Map each JHU column to a field by header keyword."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(header_cells):
        h = str(raw or "").strip().lower()
        if not h:
            continue
        if h.startswith("sponsor"):
            mapping.setdefault("sponsor", idx)
        elif h.startswith("program"):
            mapping.setdefault("program", idx)
        elif h.startswith("description"):
            mapping.setdefault("description", idx)
        elif "eligibility" in h:
            mapping.setdefault("eligibility", idx)
        elif "citizenship" in h:
            mapping.setdefault("citizenship", idx)
        elif "keyword" in h or "discipline" in h:
            mapping.setdefault("keywords", idx)
        elif "amount" in h:
            mapping.setdefault("amount", idx)
        elif "annual deadline" in h:
            mapping.setdefault("annual_deadline", idx)
        elif "deadline" in h:
            mapping.setdefault("deadline", idx)
    return mapping


def parse_worksheet(ws, cfg: dict, as_of: Optional[_dt.date] = None) -> list[dict]:
    """Turn one JHU worksheet into opportunity dicts (dependency-light; the
    caller passes an openpyxl worksheet)."""
    # Find the header row (the one containing "Sponsor").
    header_row = None
    header_cells = []
    for r in range(1, 9):
        cells = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(str(v or "").strip().lower().startswith("sponsor") for v in cells):
            header_row, header_cells = r, cells
            break
    if header_row is None:
        return []
    cols = _colmap(header_cells)
    if "program" not in cols or "sponsor" not in cols:
        return []

    def cell(r, key):
        c = cols.get(key)
        return ws.cell(r, c + 1).value if c is not None else None

    as_of = as_of or _dt.date.today()
    out: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        program = cell(r, "program")
        sponsor = cell(r, "sponsor")
        if not program or not str(program).strip():
            continue
        program = re.sub(r"\s+", " ", str(program)).strip()
        sponsor = re.sub(r"\s+", " ", str(sponsor or "")).strip()
        if cfg["drop_federal"] and sponsor and _FEDERAL_RE.search(sponsor):
            continue  # already covered by our Grants.gov catalog

        link = None
        prog_col = cols.get("program")
        if prog_col is not None:
            link = getattr(ws.cell(r, prog_col + 1), "hyperlink", None)
            link = getattr(link, "target", None)
        url = _clean_url(link, cfg["page"])

        eligibility = cell(r, "eligibility")
        citizenship = cell(r, "citizenship")
        elig_text = " ".join(
            str(x).strip() for x in (eligibility, citizenship) if x and str(x).strip()
        ) or None
        keywords = cell(r, "keywords")
        deadline_text = cell(r, "deadline") or cell(r, "annual_deadline")
        close_date, deadline_note, deadline_status = _deadline_metadata(
            deadline_text, as_of
        )

        external_id = "jhu-{}".format(
            hashlib.sha1(
                f"{sponsor}|{program}".casefold().encode()
            ).hexdigest()[:16],
        )
        out.append({
            "title": program,
            "external_id": external_id,
            "url": url,
            "agency": sponsor or "Johns Hopkins RDT list",
            "description": (str(cell(r, "description")).strip()
                            if cell(r, "description") else None),
            "eligibility_text": elig_text,
            "applicant_types": list(cfg["applicant_types"]),
            "disciplines": [str(keywords).strip()] if keywords and str(keywords).strip() else [],
            "award_ceiling": _money(cell(r, "amount")),
            "close_date": close_date,
            "deadline_note": deadline_note,
            "_deadline_status": deadline_status,
        })
    return out


class JHUFellowshipsAdapter(SourceAdapter):
    slug = "jhu-fellowships"
    display_name = "Johns Hopkins RDT fellowships list"
    source_type = "Fellowship"
    # The official pages, WordPress API, direct media URLs, and JHU-published
    # short links all return an interactive Cloudflare challenge to unattended
    # refresh clients. A complete three-workbook set is mandatory and this
    # source has no safe stale fallback, so do not represent it as healthy or
    # repeatedly fail the publication gate. Keep the parser available for
    # bounded diagnostics with --include-disabled and re-enable only after an
    # official unattended retrieval path is live-verified.
    enabled = False
    disabled_reason = (
        "official JHU workbook routes require an interactive Cloudflare "
        "challenge; no complete unattended source or safe fallback is available"
    )
    # A complete workbook set can legitimately contain zero currently open
    # rows after strict expiration filtering. Raw workbook volume is validated
    # separately in parse().
    min_records = 0
    max_records = 1500
    fallback_grace_days = 0
    retain_on_failure = False

    def __init__(self, as_of: Optional[_dt.date] = None) -> None:
        super().__init__()
        self.as_of = as_of or _dt.date.today()
        # JHU's Cloudflare edge issues a bounded bot-management cookie even on
        # successful public page requests. Reuse it for the corresponding
        # public workbook request instead of treating each URL as a new client.
        self._opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(http.cookiejar.CookieJar())
        )

    def set_context(self, context: dict) -> None:
        """Use the merge run's catalog date for all currentness decisions."""
        super().set_context(context)
        effective_as_of = self.context.get("as_of")
        if isinstance(effective_as_of, _dt.datetime):
            effective_as_of = effective_as_of.date()
        elif isinstance(effective_as_of, str):
            try:
                effective_as_of = _dt.date.fromisoformat(effective_as_of)
            except ValueError:
                effective_as_of = None
        if isinstance(effective_as_of, _dt.date):
            self.as_of = effective_as_of

    def fetch(self):
        """Download every JHU workbook or fail the source as an incomplete run.

        JHU intermittently rejects automated page requests. The page remains
        the preferred source of the current workbook URL, while the official
        bit.ly link printed on that page is a fallback. A partial download is a
        failed refresh; the no-fallback lifecycle publishes zero rather than an
        old snapshot.
        """
        results = []
        failures = []
        page_states = {}
        download_sources = {}
        snapshot_dates = {}
        for cfg in SHEETS:
            candidates = []
            page_error = None
            try:
                html = self._get(
                    cfg["page"],
                    accept="text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
                ).decode("utf-8", errors="replace")
                links = _LINK_RE.findall(html)
                candidate = next(
                    (u for u in links
                     if "bit.ly" in u.lower() and "opp" in u.lower()), None)
                candidate = candidate or next(
                    (u for u in links if u.lower().endswith(".xlsx")), None)
                if candidate:
                    candidates.append(("page_link", candidate))
                    page_states[cfg["audience"]] = "workbook_link_present"
                elif "under construction" in html.casefold():
                    page_states[cfg["audience"]] = "under_construction_no_workbook_link"
                else:
                    page_states[cfg["audience"]] = "no_workbook_link"
            except Exception as exc:  # page fallback is intentional
                page_error = f"page: {type(exc).__name__}: {exc}"
                page_states[cfg["audience"]] = "page_request_failed"
            direct = cfg.get("direct_sheet")
            if direct and all(url != direct for _, url in candidates):
                candidates.append(("official_direct", direct))
            fallback = cfg.get("fallback_sheet")
            if fallback and all(url != fallback for _, url in candidates):
                candidates.append(("official_shortlink", fallback))

            sheet_errors = []
            for candidate_kind, candidate in candidates:
                try:
                    data = self._get(
                        candidate,
                        accept=(
                            "application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet,application/octet-stream;q=0.9,"
                            "*/*;q=0.8"
                        ),
                        referer=cfg["page"],
                    )
                    if not data.startswith(b"PK"):
                        raise ValueError("response is not an XLSX/ZIP workbook")
                    results.append({
                        **cfg,
                        "data": data,
                        "sheet_url": candidate,
                        "sheet_candidate_kind": candidate_kind,
                    })
                    download_sources[cfg["audience"]] = candidate_kind
                    if (
                        cfg.get("snapshot_date")
                        and candidate in {
                            cfg.get("direct_sheet"),
                            cfg.get("fallback_sheet"),
                        }
                    ):
                        snapshot_dates[cfg["audience"]] = cfg["snapshot_date"]
                    break
                except Exception as exc:
                    sheet_errors.append(
                        f"{candidate_kind} {candidate}: {type(exc).__name__}: {exc}"
                    )
            else:
                details = [page_error, *sheet_errors]
                failures.append(
                    f"{cfg['audience']} ({'; '.join(item for item in details if item)})"
                )
        self.diagnostics = {
            "expected_audiences": [cfg["audience"] for cfg in SHEETS],
            "downloaded_audiences": [cfg["audience"] for cfg in results],
            "download_failures": failures,
            "page_states": page_states,
            "download_sources_by_audience": download_sources,
            "snapshot_dates_by_audience": snapshot_dates,
        }
        if failures or len(results) != len(SHEETS):
            joined = " | ".join(failures)
            if "HTTP 403" in joined or "HTTP Error 403" in joined:
                self.diagnostics["failure_class"] = "request_network"
                self.diagnostics["failure_reason"] = "http_403_access_challenge"
            elif "response is not an XLSX/ZIP workbook" in joined:
                self.diagnostics["failure_class"] = "upstream_response_change"
                self.diagnostics["failure_reason"] = "non_workbook_response"
            else:
                self.diagnostics["failure_class"] = "request_network"
                self.diagnostics["failure_reason"] = "workbook_request_failed"
            raise RuntimeError("Incomplete JHU workbook refresh: " + " | ".join(failures))
        if snapshot_dates:
            oldest_snapshot = min(
                _dt.date.fromisoformat(value) for value in snapshot_dates.values()
            )
            snapshot_age_days = (self.as_of - oldest_snapshot).days
            self.diagnostics.update({
                "source_state": "bounded_official_snapshot",
                "source_snapshot_at": oldest_snapshot.isoformat(),
                "source_snapshot_age_days": snapshot_age_days,
                "source_snapshot_max_age_days": PINNED_WORKBOOK_MAX_AGE_DAYS,
            })
            if snapshot_age_days < 0:
                self.diagnostics["failure_class"] = "upstream_response_change"
                self.diagnostics["failure_reason"] = "pinned_workbook_newer_than_catalog"
                raise RuntimeError(
                    "Incomplete JHU workbook refresh: official fallback snapshot "
                    f"is newer than catalog date {self.as_of.isoformat()}"
                )
            if snapshot_age_days > PINNED_WORKBOOK_MAX_AGE_DAYS:
                self.diagnostics["failure_class"] = "upstream_response_change"
                self.diagnostics["failure_reason"] = "pinned_workbook_expired"
                raise RuntimeError(
                    "Incomplete JHU workbook refresh: official fallback snapshot "
                    f"is {snapshot_age_days} days old (maximum "
                    f"{PINNED_WORKBOOK_MAX_AGE_DAYS})"
                )
        else:
            self.diagnostics["source_state"] = "live_page_workbooks"
        return results

    def _get(
        self,
        url: str,
        *,
        accept: str = "*/*",
        referer: Optional[str] = None,
    ) -> bytes:
        last_error = None
        for attempt in range(FETCH_ATTEMPTS):
            try:
                headers = {
                    "User-Agent": USER_AGENT,
                    "Accept": accept,
                    "Accept-Language": "en-US,en;q=0.8",
                }
                if referer:
                    headers["Referer"] = referer
                req = urllib.request.Request(url, headers=headers)
                with self._opener.open(req, timeout=60) as resp:
                    return resp.read(16 * 1024 * 1024)  # 16 MB cap
            except urllib.error.HTTPError as exc:
                mitigation = exc.headers.get("Cf-Mitigated") if exc.headers else None
                suffix = f"; cf-mitigated={mitigation}" if mitigation else ""
                last_error = RuntimeError(
                    f"HTTP Error {exc.code}: {exc.reason}{suffix}"
                )
                if attempt + 1 < FETCH_ATTEMPTS:
                    time.sleep(attempt + 1)
            except Exception as exc:
                last_error = exc
                if attempt + 1 < FETCH_ATTEMPTS:
                    time.sleep(attempt + 1)
        raise RuntimeError(
            f"request failed after {FETCH_ATTEMPTS} attempts: {last_error}"
        ) from last_error

    def parse(self, payload) -> Iterable[CanonicalOpportunity]:
        import openpyxl
        merged: dict[str, dict] = {}
        raw_counts = {}
        current_counts = {}
        dropped_counts = {}
        for cfg in payload or []:
            wb = openpyxl.load_workbook(io.BytesIO(cfg["data"]), data_only=True)
            items = parse_worksheet(wb.active, cfg, self.as_of)
            raw_counts[cfg["audience"]] = len(items)
            if len(items) < MIN_ROWS_PER_AUDIENCE:
                self.diagnostics = {
                    **getattr(self, "diagnostics", {}),
                    "raw_rows_by_audience": raw_counts,
                    "failure_class": "health_bound_violation",
                    "failure_reason": "audience_row_floor",
                }
                raise ValueError(
                    f"JHU {cfg['audience']} workbook yielded only {len(items)} rows; "
                    f"expected at least {MIN_ROWS_PER_AUDIENCE}."
                )
            drop_reasons: Counter = Counter()
            current = 0
            for item in items:
                status = item.pop("_deadline_status")
                if status not in {"current", "rolling"}:
                    drop_reasons[status] += 1
                    continue
                current += 1
                external_id = item["external_id"]
                existing = merged.get(external_id)
                if existing:
                    for field in ("applicant_types", "disciplines"):
                        existing[field] = list(dict.fromkeys(
                            [*(existing.get(field) or []), *(item.get(field) or [])]
                        ))
                    continue
                merged[external_id] = item
            current_counts[cfg["audience"]] = current
            dropped_counts[cfg["audience"]] = dict(sorted(drop_reasons.items()))
        expected = {cfg["audience"] for cfg in SHEETS}
        if set(raw_counts) != expected:
            raise ValueError(
                "JHU parse did not receive every audience: "
                + ", ".join(sorted(expected - set(raw_counts)))
            )
        parsed = [self._to_canonical(item) for item in merged.values()]
        self.diagnostics = {
            **getattr(self, "diagnostics", {}),
            "raw_rows_by_audience": raw_counts,
            "current_rows_by_audience": current_counts,
            "dropped_deadlines_by_audience": dropped_counts,
            "deduplicated_current_rows": sum(current_counts.values()) - len(parsed),
            "parsed_records": len(parsed),
        }
        return parsed

    def parse_file(self, path: str, cfg: dict) -> list[CanonicalOpportunity]:
        """Offline test helper: parse a local .xlsx against a sheet config."""
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        return [
            self._to_canonical({k: v for k, v in item.items() if not k.startswith("_")})
            for item in parse_worksheet(wb.active, cfg, self.as_of)
            if item.get("_deadline_status") in {"current", "rolling"}
        ]

    @staticmethod
    def _to_canonical(item: dict) -> CanonicalOpportunity:
        return CanonicalOpportunity(
            title=item["title"],
            external_id=item.get("external_id"),
            url=item.get("url"),
            agency=item.get("agency"),
            description=item.get("description"),
            eligibility_text=item.get("eligibility_text"),
            applicant_types=item.get("applicant_types") or [],
            disciplines=item.get("disciplines") or [],
            award_ceiling=item.get("award_ceiling"),
            close_date=item.get("close_date"),
            deadline_note=item.get("deadline_note"),
        )


register(JHUFellowshipsAdapter())
