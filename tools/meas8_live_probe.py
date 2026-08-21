"""MEAS-8 live source capture and pre-Cov4 production probe.

The outcome-blind frame is already committed.  This runner re-derives current
sources for Arm A, preserves extracted text under ``.work/meas8`` for independent
DEC-11 truth labeling, and invokes the existing production source selection,
segmentation, and record builder.  It deliberately does not call Cov4: MEAS-8
establishes human truth and deterministic candidates before spending model calls.
"""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
import sys
import zipfile
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts import subtopic_records, subtopic_referenced, subtopic_sources  # noqa: E402
from scripts.extract_document_evidence import (  # noqa: E402
    download_document,
    extract_containers,
    referenced_fetch,
    source_for_record,
    subtopic_fields,
)
from scripts.pull_grants import collect_attachments, fetch_detail  # noqa: E402
from tools import p7_frame  # noqa: E402


ROOT = Path(__file__).resolve().parents[1]
FRAME = ROOT / "evaluation" / "meas8_frame.json"
DEFAULT_OUT = ROOT / ".work" / "meas8" / "arm_a_probe.json"
DEFAULT_ARM_B_OUT = ROOT / ".work" / "meas8" / "arm_b_parent_probe.json"
DEFAULT_COV4_OUT = ROOT / ".work" / "meas8" / "arm_a_cov4.json"
AS_OF = "2026-08-20"


def _safe(value):
    return re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "source"))[:100]


def _docx_text(content):
    with zipfile.ZipFile(BytesIO(content)) as archive:
        raw = archive.read("word/document.xml").decode("utf-8", errors="replace")
    raw = re.sub(r"</w:p>", "\n", raw)
    raw = re.sub(r"<w:tab[^>]*/>", "\t", raw)
    raw = re.sub(r"<[^>]+>", "", raw)
    return html.unescape(raw)


def _xlsx_text(content):
    from openpyxl import load_workbook

    book = load_workbook(BytesIO(content), read_only=True, data_only=True)
    lines = []
    for sheet in book.worksheets:
        lines.append(f"===== SHEET: {sheet.title} =====")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if values:
                lines.append("\t".join(values))
    return "\n".join(lines)


def truth_text(content, content_type, name, url):
    suffix = f"{name or ''} {url or ''}".casefold()
    if ".docx" in suffix or content.startswith(b"PK\x03\x04") and b"word/" in content[:5000]:
        return _docx_text(content), {"method": "measurement_only_docx", "content_kind": "docx"}
    if ".xlsx" in suffix:
        return _xlsx_text(content), {"method": "measurement_only_xlsx", "content_kind": "xlsx"}
    containers, extraction = extract_containers(content, content_type, name, url)
    lines = []
    for container in containers:
        label = container.get("page") or container.get("section") or "Official notice"
        lines.append(f"===== {label} =====")
        lines.append(container.get("text") or "")
    return "\n".join(lines), extraction


class LiveCache:
    def __init__(self):
        self.downloads = {}
        self.details = {}

    def download(self, url):
        if url not in self.downloads:
            self.downloads[url] = download_document(url)
        return self.downloads[url]

    def detail(self, opportunity_id):
        key = str(opportunity_id)
        if key not in self.details:
            self.details[key] = fetch_detail(key)
        return self.details[key]


def _source_list(record, live_detail, primary):
    sources = []
    if primary:
        sources.append({"url": primary.get("url"), "name": primary.get("name"), "kind": primary.get("kind")})
    if live_detail is not None:
        data = live_detail.get("data", live_detail)
        for item in collect_attachments(data):
            sources.append({
                "url": item.get("download_url"),
                "name": item.get("file_name"),
                "kind": "grants_gov_attachment",
                "attachment_id": item.get("id"),
                "mime_type_at_source": item.get("mime_type"),
                "size_bytes_at_source": item.get("size_bytes"),
            })
    deduped = []
    seen = set()
    for source in sources:
        if not source.get("url") or source["url"] in seen:
            continue
        seen.add(source["url"])
        deduped.append(source)
    return deduped


def _capture_source(record_id, ordinal, source, cache, workdir):
    row = dict(source)
    try:
        response = cache.download(source["url"])
        content = response.get("content") or b""
        digest = hashlib.sha256(content).hexdigest() if content else None
        row.update({
            "status": "fetched",
            "final_url": response.get("url") or source["url"],
            "content_type": response.get("content_type"),
            "sha256": digest,
            "bytes": len(content),
        })
        text, extraction = truth_text(
            content, response.get("content_type"), source.get("name"), row["final_url"]
        )
        text_name = f"{_safe(record_id)}_{ordinal:02d}_{_safe(source.get('name') or source.get('kind'))}_{digest[:12]}.txt"
        text_path = workdir / text_name
        text_path.write_text(text, encoding="utf-8", newline="\n")
        row["truth_text_path"] = str(text_path.resolve().relative_to(ROOT)).replace("\\", "/")
        row["truth_text_characters"] = len(text)
        row["extraction"] = extraction
    except Exception as exc:  # report each layer independently
        row.update({"status": "failed", "error_type": type(exc).__name__, "error": str(exc)})
    return row


def _segmentation_summary(result, document, diagnostics, built, referenced_diag):
    return {
        "method": result.method if result else None,
        "family": result.family if result else None,
        "confidence": result.confidence if result else None,
        "reason": result.reason if result else None,
        "chosen_document": document,
        "diagnostics": diagnostics,
        "referenced_diagnostics": referenced_diag,
        "candidate_count": len(built),
        "candidate_titles": [row["title"] for row in built],
        "candidates": [
            {
                "subtopic_id": row.get("subtopic_id"),
                "title": row.get("title"),
                "subtopic_code": row.get("subtopic_code"),
                "provenance": row.get("subtopic_source"),
                "confidence": row.get("confidence"),
                "summary": row.get("summary"),
                "page_start": row.get("page_start"),
            }
            for row in built
        ],
    }


def probe_record(record, workdir):
    record_id = str(record["opportunity_id"])
    print(f"probe {record_id} {record.get('opportunity_number') or ''}", flush=True)
    cache = LiveCache()
    detail = None
    detail_error = None
    if record.get("source") == "Grants.gov" and record_id.isdigit():
        try:
            detail = cache.detail(record_id)
        except Exception as exc:
            detail_error = {"error_type": type(exc).__name__, "error": str(exc)}

    designated = source_for_record(record)
    primary = designated or subtopic_sources.subtopic_only_primary(record)
    sources = _source_list(record, detail, primary)
    captured = [
        _capture_source(record_id, ordinal, source, cache, workdir)
        for ordinal, source in enumerate(sources, start=1)
    ]

    primary_content = None
    primary_document = None
    if primary and primary.get("url") in cache.downloads:
        response = cache.downloads[primary["url"]]
        content = response.get("content") or b""
        if content:
            primary_content = content
            primary_document = {
                "url": response.get("url") or primary["url"],
                "name": primary.get("name"),
                "content_type": response.get("content_type"),
                "sha256": hashlib.sha256(content).hexdigest(),
                "source_kind": primary.get("kind"),
            }

    referenced_result, referenced_document, referenced_diag = subtopic_referenced.first_refusal(
        record, fetch=referenced_fetch
    )
    if referenced_result is not None:
        result, chosen, diagnostics = referenced_result, referenced_document, {"attempts": []}
        provenance = subtopic_records.REFERENCED
    else:
        result, chosen, diagnostics = subtopic_sources.best_segmentation(
            record,
            primary_content,
            primary_document,
            extract_containers=extract_containers,
            download=cache.download,
            detail_fetcher=cache.detail,
            collector=collect_attachments,
            parent_deadline=record.get("close_date"),
        )
        provenance = subtopic_records.INFERRED
    built = subtopic_records.build_records(
        record,
        result,
        document=chosen or primary_document,
        as_of=AS_OF,
        provenance=provenance,
    )
    live_attachments = []
    if detail is not None:
        live_attachments = collect_attachments(detail.get("data", detail))
    return {
        "opportunity_id": record_id,
        "opportunity_number": record.get("opportunity_number"),
        "title": record.get("title"),
        "agency": record.get("agency"),
        "catalog_status": record.get("status"),
        "catalog_actionability": record.get("actionability_status"),
        "live_detail": {
            "status": "fetched" if detail is not None else ("failed" if detail_error else "not_applicable"),
            "error": detail_error,
            "attachment_count": len(live_attachments),
            "attachments": live_attachments,
        },
        "sources": captured,
        "production_pre_cov4": _segmentation_summary(
            result, chosen or primary_document, diagnostics, built, referenced_diag
        ),
    }


def run(out_path=DEFAULT_OUT):
    frame = json.loads(FRAME.read_text(encoding="utf-8"))
    catalog = p7_frame.load_catalog()
    by_id = {str(row["opportunity_id"]): row for row in catalog["opportunities"]}
    selected = [
        row["opportunity_id"]
        for stratum in frame["arm_a"]["strata"]
        for row in stratum["records"]
    ]
    workdir = out_path.parent / "truth_text"
    workdir.mkdir(parents=True, exist_ok=True)
    rows = []
    for record_id in selected:
        try:
            rows.append(probe_record(by_id[record_id], workdir))
        except Exception as exc:
            rows.append({
                "opportunity_id": record_id,
                "status": "processing_error",
                "error_type": type(exc).__name__,
                "error": str(exc),
            })
    payload = {
        "schema_version": 1,
        "frame_commit": "16b765f",
        "phase": "Arm A live source capture and deterministic production through build_records; Cov4 not run",
        "records": rows,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return payload


def run_records(record_ids, out_path=DEFAULT_ARM_B_OUT):
    """Probe an explicit, already-frozen record list without changing the frame."""
    catalog = p7_frame.load_catalog()
    by_id = {str(row["opportunity_id"]): row for row in catalog["opportunities"]}
    workdir = out_path.parent / "truth_text"
    workdir.mkdir(parents=True, exist_ok=True)
    missing = [str(record_id) for record_id in record_ids if str(record_id) not in by_id]
    if missing:
        raise KeyError(f"record ids absent from the frozen catalog: {', '.join(missing)}")
    rows = []
    for record_id in record_ids:
        try:
            rows.append(probe_record(by_id[str(record_id)], workdir))
        except Exception as exc:
            rows.append({
                "opportunity_id": str(record_id),
                "status": "processing_error",
                "error_type": type(exc).__name__,
                "error": str(exc),
            })
    payload = {
        "schema_version": 1,
        "frame_commit": "16b765f",
        "phase": "Arm B frozen-parent live source capture and deterministic production through build_records; Cov4 not run",
        "records": rows,
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return payload


def run_cov4(record_id, out_path=DEFAULT_COV4_OUT):
    """Run the exact production call site for one truth-positive candidate set."""
    catalog = p7_frame.load_catalog()
    by_id = {str(row["opportunity_id"]): row for row in catalog["opportunities"]}
    record = by_id[str(record_id)]
    source = source_for_record(record) or subtopic_sources.subtopic_only_primary(record)
    if not source:
        raise RuntimeError(f"{record_id} has no production primary source")
    response = download_document(source["url"])
    content = response.get("content") or b""
    document = {
        "url": response.get("url") or source["url"],
        "name": source.get("name"),
        "content_type": response.get("content_type"),
        "sha256": hashlib.sha256(content).hexdigest(),
        "source_kind": source.get("kind"),
    }
    containers, extraction = extract_containers(
        content, document["content_type"], document["name"], document["url"]
    )
    fields = subtopic_fields(
        record,
        content,
        containers,
        document,
        f"{AS_OF}T00:00:00Z",
        True,
    )
    payload = {
        "schema_version": 1,
        "frame_commit": "16b765f",
        "opportunity_id": str(record_id),
        "document": document,
        "extraction": extraction,
        "subtopic_method": fields.get("subtopic_method"),
        "subtopic_reason": fields.get("subtopic_reason"),
        "subtopic_attempts": fields.get("subtopic_attempts"),
        "cov4": fields.get("subtopic_cov4"),
        "subtopics": [
            {
                "title": row.get("title"),
                "subtopic_code": row.get("subtopic_code"),
                "provenance": row.get("subtopic_source"),
                "confidence": row.get("confidence"),
                "review_required": row.get("review_required"),
                "publishable": subtopic_records.is_publishable(row),
                "page_start": row.get("page_start"),
            }
            for row in fields.get("subtopics", [])
        ],
    }
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return payload


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--records", nargs="+")
    parser.add_argument("--cov4-record")
    parser.add_argument("--cov4-out", type=Path, default=DEFAULT_COV4_OUT)
    args = parser.parse_args(argv)
    if args.cov4_record:
        payload = run_cov4(args.cov4_record, args.cov4_out)
        print(f"wrote {args.cov4_out}", flush=True)
        print(
            f"classifier_calls {(payload.get('cov4') or {}).get('classifier_calls')} "
            f"subtopics {len(payload['subtopics'])}",
            flush=True,
        )
        return
    if args.records:
        payload = run_records(args.records, args.out)
        print(f"wrote {args.out}", flush=True)
        print(f"records {len(payload['records'])}", flush=True)
        return
    payload = run(args.out)
    print(f"wrote {args.out}", flush=True)
    print(f"records {len(payload['records'])}", flush=True)


if __name__ == "__main__":
    main()
